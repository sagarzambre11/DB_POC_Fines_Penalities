"""
app/reporter.py
---------------
Step 5: Two-layer gap analysis report generation and Excel export.

Produces:
  - Policy Gap DataFrame   (Layer 1 — primary, shift-left signal)
  - Control Gap DataFrame  (Layer 2 — operational signal)
  - Stakeholder Signals DataFrame
  - Unaddressed Findings DataFrame
  - Downloadable 6-sheet Excel workbook
"""

import io
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.comparator import (
    get_policy_gap_rows,
    get_control_gap_rows,
    get_stakeholder_signal_rows,
    get_unaddressed_findings_rows,
    get_overall_assessment,
)

# ---------------------------------------------------------------------------
# Colour maps
# ---------------------------------------------------------------------------

POLICY_FILL = {
    "Covered":               "C6EFCE",
    "Partially Covered":     "FFEB9C",
    "Potential Gap":         "FFC7CE",
    "Insufficient Evidence": "D9D9D9",
}
POLICY_FONT = {
    "Covered":               "276221",
    "Partially Covered":     "9C5700",
    "Potential Gap":         "9C0006",
    "Insufficient Evidence": "595959",
}
CONTROL_FILL = {
    "Covered":               "C6EFCE",
    "Partially Covered":     "FFEB9C",
    "Policy-Only Coverage":  "BDD7EE",
    "Potential Gap":         "FFC7CE",
    "Insufficient Evidence": "D9D9D9",
}
CONTROL_FONT = {
    "Covered":               "276221",
    "Partially Covered":     "9C5700",
    "Policy-Only Coverage":  "1F4E79",
    "Potential Gap":         "9C0006",
    "Insufficient Evidence": "595959",
}
SEVERITY_FILL = {
    "Critical": "FF0000",
    "High":     "FFC7CE",
    "Medium":   "FFEB9C",
    "Low":      "C6EFCE",
}
SEVERITY_FONT = {
    "Critical": "FFFFFF",
    "High":     "9C0006",
    "Medium":   "9C5700",
    "Low":      "276221",
}
PRIORITY_FILL = {
    "High":   "FFC7CE",
    "Medium": "FFEB9C",
    "Low":    "C6EFCE",
}
RISK_FILL = {
    "Critical": "FF0000",
    "High":     "FFC7CE",
    "Medium":   "FFEB9C",
    "Low":      "C6EFCE",
}


# ---------------------------------------------------------------------------
# DataFrame builders
# ---------------------------------------------------------------------------

def build_policy_gap_dataframe(comparison: dict) -> pd.DataFrame:
    """Build Layer 1 — Policy Gap DataFrame."""
    rows = get_policy_gap_rows(comparison)
    cols = [
        "ID", "Name", "Domain", "Policy Owner",
        "Policy Coverage", "Rationale", "Enforcement Evidence",
        "Shift Left Signal", "Recommended Action",
        "Gap Severity", "Related Themes",
    ]
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def build_control_gap_dataframe(comparison: dict) -> pd.DataFrame:
    """Build Layer 2 — Control Gap DataFrame."""
    rows = get_control_gap_rows(comparison)
    cols = [
        "ID", "Name", "Domain", "Control Owner",
        "Control Coverage", "Rationale", "Enforcement Evidence",
        "Recommended Action", "Gap Severity", "Related Themes",
    ]
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def build_stakeholder_signals_dataframe(comparison: dict) -> pd.DataFrame:
    """Build Stakeholder Signals DataFrame."""
    rows = get_stakeholder_signal_rows(comparison)
    cols = ["ID", "Name", "Stakeholder", "Signal", "Priority", "Gap Severity"]
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def build_unaddressed_findings_dataframe(comparison: dict) -> pd.DataFrame:
    """Build Unaddressed Findings DataFrame."""
    rows = get_unaddressed_findings_rows(comparison)
    cols = [
        "Enforcement Theme", "Risk Implication",
        "Suggested Policy", "Suggested Control", "Suggested Owner",
    ]
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def build_summary_dataframe(comparison: dict) -> pd.DataFrame:
    """Build a summary metrics DataFrame from overall_assessment."""
    a = get_overall_assessment(comparison)
    pl = a.get("policy_layer_summary", {})
    cl = a.get("control_layer_summary", {})
    rows = [
        ("Total Items Assessed", a.get("total_assessed", 0)),
        ("--- POLICY LAYER ---", ""),
        ("  Policy: Covered", pl.get("covered", 0)),
        ("  Policy: Partially Covered", pl.get("partially_covered", 0)),
        ("  Policy: Potential Gap", pl.get("potential_gap", 0)),
        ("  Policy: Insufficient Evidence", pl.get("insufficient_evidence", 0)),
        ("--- CONTROL LAYER ---", ""),
        ("  Control: Covered", cl.get("covered", 0)),
        ("  Control: Partially Covered", cl.get("partially_covered", 0)),
        ("  Control: Policy-Only Coverage", cl.get("policy_only", 0)),
        ("  Control: Potential Gap", cl.get("potential_gap", 0)),
        ("  Control: Insufficient Evidence", cl.get("insufficient_evidence", 0)),
        ("Overall Risk Rating", a.get("overall_risk_rating", "N/A")),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


# ---------------------------------------------------------------------------
# Excel helper utilities
# ---------------------------------------------------------------------------

def _header_style(ws, row: int, num_cols: int) -> None:
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border(cell) -> None:
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws, min_w: int = 12, max_w: int = 55) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(c.value)) for c in col if c.value), default=0
        )
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 4, max_w))


def _colour_cell(cell, fill_hex: str, font_hex: str = "000000", bold: bool = True) -> None:
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    cell.font = Font(color=font_hex, bold=bold)


def _write_df_to_sheet(
    ws,
    df: pd.DataFrame,
    classification_col: str = None,
    fill_map: dict = None,
    font_map: dict = None,
    severity_col: str = None,
    priority_col: str = None,
    row_height: int = 55,
) -> None:
    """Write a DataFrame to a worksheet with optional colour coding."""
    headers = list(df.columns)
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    class_idx = (headers.index(classification_col) + 1) if classification_col and classification_col in headers else None
    sev_idx = (headers.index(severity_col) + 1) if severity_col and severity_col in headers else None
    pri_idx = (headers.index(priority_col) + 1) if priority_col and priority_col in headers else None

    for _, row in df.iterrows():
        ws.append(list(row))
        rn = ws.max_row
        classification = str(row.get(classification_col, "")) if classification_col else ""
        severity = str(row.get(severity_col, "")) if severity_col else ""
        priority = str(row.get(priority_col, "")) if priority_col else ""

        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=rn, column=ci)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            _border(cell)

        if class_idx and fill_map:
            bg = fill_map.get(classification, "FFFFFF")
            fg = (font_map or {}).get(classification, "000000")
            _colour_cell(ws.cell(row=rn, column=class_idx), bg, fg)

        if sev_idx:
            bg = SEVERITY_FILL.get(severity, "FFFFFF")
            fg = SEVERITY_FONT.get(severity, "000000")
            _colour_cell(ws.cell(row=rn, column=sev_idx), bg, fg)

        if pri_idx:
            bg = PRIORITY_FILL.get(priority, "FFFFFF")
            _colour_cell(ws.cell(row=rn, column=pri_idx), bg)

        ws.row_dimensions[rn].height = row_height

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Main Excel report generator
# ---------------------------------------------------------------------------

def generate_excel_report(
    comparison: dict,
    extracted_enforcement: dict,
    inventory: list[dict],
) -> bytes:
    """
    Generate a 6-sheet Excel workbook with the full two-layer gap analysis.

    Sheets:
      1. Summary              — Overall assessment, shift-left headline, risk rating
      2. Policy Gap Analysis  — Layer 1: policy coverage per inventory item
      3. Control Gap Analysis — Layer 2: control coverage per inventory item
      4. Stakeholder Signals  — Who needs to act, what, and with what priority
      5. Unaddressed Findings — Enforcement themes with no matching policy or control
      6. Enforcement Data     — Extracted enforcement intelligence
      (GRC Inventory appended as sheet 7 for reference)

    Args:
        comparison:            Dict from comparator.compare_findings_to_inventory().
        extracted_enforcement: Dict from extractor.extract_enforcement_data().
        inventory:             List of control dicts from inventory.load_inventory().

    Returns:
        Excel file as bytes.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    assessment = get_overall_assessment(comparison)
    entity = extracted_enforcement.get("regulated_entity", {})
    action = extracted_enforcement.get("enforcement_action", {})
    regulator = extracted_enforcement.get("regulator", {})

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")

    ws_s.merge_cells("A1:C1")
    ws_s["A1"].value = "Regulatory Enforcement Intelligence — Gap Analysis Report"
    ws_s["A1"].font = Font(bold=True, size=15, color="1F4E79")
    ws_s["A1"].alignment = Alignment(horizontal="center")
    ws_s.row_dimensions[1].height = 32

    # Shift-left headline banner
    headline = assessment.get("shift_left_headline", "")
    if headline:
        ws_s.merge_cells("A2:C2")
        ws_s["A2"].value = f"⚡ {headline}"
        ws_s["A2"].font = Font(bold=True, size=12, color="9C0006")
        ws_s["A2"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws_s["A2"].alignment = Alignment(horizontal="center", wrap_text=True)
        ws_s.row_dimensions[2].height = 28

    # Metadata
    meta = [
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Regulated Entity", entity.get("name", "N/A")),
        ("Regulator", f"{regulator.get('name', 'N/A')} ({regulator.get('abbreviation', '')})"),
        ("Jurisdiction", extracted_enforcement.get("jurisdiction", "N/A")),
        ("Notice Date", action.get("notice_date", "N/A")),
        ("Penalty", f"{action.get('penalty_currency', '')} {action.get('penalty_amount', 'N/A')}"),
        ("Reference", action.get("reference_number", "N/A")),
        ("Domain(s)", ", ".join(extracted_enforcement.get("regulatory_domain", []))),
    ]
    for i, (lbl, val) in enumerate(meta, start=3):
        ws_s.cell(row=i, column=1, value=lbl).font = Font(bold=True)
        ws_s.cell(row=i, column=2, value=str(val))

    ws_s.append([])

    # Policy layer summary
    pl = assessment.get("policy_layer_summary", {})
    cl = assessment.get("control_layer_summary", {})

    ws_s.append(["Layer", "Classification", "Count"])
    _header_style(ws_s, ws_s.max_row, 3)

    policy_rows = [
        ("Policy Layer", "Covered", pl.get("covered", 0), "C6EFCE", "276221"),
        ("Policy Layer", "Partially Covered", pl.get("partially_covered", 0), "FFEB9C", "9C5700"),
        ("Policy Layer", "Potential Gap", pl.get("potential_gap", 0), "FFC7CE", "9C0006"),
        ("Policy Layer", "Insufficient Evidence", pl.get("insufficient_evidence", 0), "D9D9D9", "595959"),
        ("Control Layer", "Covered", cl.get("covered", 0), "C6EFCE", "276221"),
        ("Control Layer", "Partially Covered", cl.get("partially_covered", 0), "FFEB9C", "9C5700"),
        ("Control Layer", "Policy-Only Coverage", cl.get("policy_only", 0), "BDD7EE", "1F4E79"),
        ("Control Layer", "Potential Gap", cl.get("potential_gap", 0), "FFC7CE", "9C0006"),
        ("Control Layer", "Insufficient Evidence", cl.get("insufficient_evidence", 0), "D9D9D9", "595959"),
    ]
    for layer, label, count, bg, fg in policy_rows:
        rn = ws_s.max_row + 1
        ws_s.cell(row=rn, column=1, value=layer)
        ws_s.cell(row=rn, column=2, value=label)
        ws_s.cell(row=rn, column=3, value=count)
        for ci in [1, 2, 3]:
            cell = ws_s.cell(row=rn, column=ci)
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.font = Font(color=fg, bold=True)
            _border(cell)

    ws_s.append([])
    risk = assessment.get("overall_risk_rating", "N/A")
    ws_s.append(["Overall Risk Rating", risk])
    risk_row = ws_s.max_row
    risk_bg = RISK_FILL.get(risk, "FFFFFF")
    for ci in [1, 2]:
        cell = ws_s.cell(row=risk_row, column=ci)
        cell.fill = PatternFill(start_color=risk_bg, end_color=risk_bg, fill_type="solid")
        cell.font = Font(bold=True, size=12)

    ws_s.append([])
    ws_s.append(["Executive Summary"])
    ws_s.cell(row=ws_s.max_row, column=1).font = Font(bold=True)
    exec_row = ws_s.max_row + 1
    ws_s.cell(row=exec_row, column=1, value=assessment.get("executive_summary", ""))
    ws_s.cell(row=exec_row, column=1).alignment = Alignment(wrap_text=True)
    ws_s.merge_cells(start_row=exec_row, start_column=1, end_row=exec_row, end_column=3)
    ws_s.row_dimensions[exec_row].height = 90
    _auto_width(ws_s)

    # ── Sheet 2: Policy Gap Analysis ─────────────────────────────────────────
    ws_p = wb.create_sheet("Policy Gap Analysis")
    policy_df = build_policy_gap_dataframe(comparison)
    _write_df_to_sheet(
        ws_p, policy_df,
        classification_col="Policy Coverage",
        fill_map=POLICY_FILL, font_map=POLICY_FONT,
        severity_col="Gap Severity",
    )

    # ── Sheet 3: Control Gap Analysis ────────────────────────────────────────
    ws_c = wb.create_sheet("Control Gap Analysis")
    control_df = build_control_gap_dataframe(comparison)
    _write_df_to_sheet(
        ws_c, control_df,
        classification_col="Control Coverage",
        fill_map=CONTROL_FILL, font_map=CONTROL_FONT,
        severity_col="Gap Severity",
    )

    # ── Sheet 4: Stakeholder Signals ─────────────────────────────────────────
    ws_ss = wb.create_sheet("Stakeholder Signals")
    signals_df = build_stakeholder_signals_dataframe(comparison)
    if not signals_df.empty:
        _write_df_to_sheet(
            ws_ss, signals_df,
            priority_col="Priority",
            severity_col="Gap Severity",
        )
    else:
        ws_ss["A1"] = "No stakeholder signals generated."

    # ── Sheet 5: Unaddressed Findings ────────────────────────────────────────
    ws_u = wb.create_sheet("Unaddressed Findings")
    unaddressed_df = build_unaddressed_findings_dataframe(comparison)
    if not unaddressed_df.empty:
        headers_u = list(unaddressed_df.columns)
        ws_u.append(headers_u)
        _header_style(ws_u, 1, len(headers_u))
        for _, row in unaddressed_df.iterrows():
            ws_u.append(list(row))
            rn = ws_u.max_row
            for ci in range(1, len(headers_u) + 1):
                cell = ws_u.cell(row=rn, column=ci)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                _border(cell)
            ws_u.row_dimensions[rn].height = 60
        ws_u.freeze_panes = "A2"
        _auto_width(ws_u)
    else:
        ws_u["A1"] = "No unaddressed findings — all enforcement themes are covered by the inventory."

    # ── Sheet 6: Enforcement Data ────────────────────────────────────────────
    ws_e = wb.create_sheet("Enforcement Data")
    ws_e.append(["Field", "Value"])
    _header_style(ws_e, 1, 2)

    enf_rows = [
        ("Regulator", regulator.get("name", "")),
        ("Regulator Abbreviation", regulator.get("abbreviation", "")),
        ("Regulator Country", regulator.get("country", "")),
        ("Jurisdiction", extracted_enforcement.get("jurisdiction", "")),
        ("Regulated Entity", entity.get("name", "")),
        ("Entity Abbreviation", entity.get("abbreviation", "")),
        ("Entity Type", entity.get("entity_type", "")),
        ("Business Context", entity.get("business_context", "")),
        ("Action Type", action.get("action_type", "")),
        ("Penalty Amount", action.get("penalty_amount", "")),
        ("Penalty Currency", action.get("penalty_currency", "")),
        ("Legal Basis", action.get("legal_basis", "")),
        ("Settlement Discount %", action.get("settlement_discount", {}).get("percentage", "")),
        ("Pre-Discount Penalty", action.get("settlement_discount", {}).get("pre_discount_penalty", "")),
        ("Notice Date", action.get("notice_date", "")),
        ("Reference Number", action.get("reference_number", "")),
        ("Remedial Outcome", action.get("additional_remedial_outcome", "")),
        ("Scenario Description", extracted_enforcement.get("scenario_description", "")),
        ("Regulatory Domains", ", ".join(extracted_enforcement.get("regulatory_domain", []))),
        (
            "Misconduct Themes",
            "\n".join(extracted_enforcement.get("misconduct_control_failure_themes", [])),
        ),
        (
            "Confidence Score",
            str(extracted_enforcement.get("confidence_score", {}).get("score", "")),
        ),
    ]
    for label, value in enf_rows:
        rn = ws_e.max_row + 1
        ws_e.cell(row=rn, column=1, value=label).font = Font(bold=True)
        cell_v = ws_e.cell(row=rn, column=2, value=str(value))
        cell_v.alignment = Alignment(wrap_text=True, vertical="top")
        _border(ws_e.cell(row=rn, column=1))
        _border(ws_e.cell(row=rn, column=2))
    _auto_width(ws_e)

    # ── Sheet 7: GRC Inventory ───────────────────────────────────────────────
    ws_inv = wb.create_sheet("GRC Inventory")
    if inventory:
        inv_headers = list(inventory[0].keys())
        ws_inv.append(inv_headers)
        _header_style(ws_inv, 1, len(inv_headers))
        for ctrl in inventory:
            ws_inv.append([ctrl.get(h, "") for h in inv_headers])
            rn = ws_inv.max_row
            for ci in range(1, len(inv_headers) + 1):
                cell = ws_inv.cell(row=rn, column=ci)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                _border(cell)
            ws_inv.row_dimensions[rn].height = 40
        ws_inv.freeze_panes = "A2"
        _auto_width(ws_inv)

    # ── Save to bytes ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def get_report_filename(extracted_enforcement: dict) -> str:
    """Generate a descriptive filename for the Excel report."""
    entity = extracted_enforcement.get("regulated_entity", {})
    regulator = extracted_enforcement.get("regulator", {})
    abbrev = (
        entity.get("abbreviation")
        or entity.get("name", "entity")
    ).replace(" ", "_").replace("/", "-")[:15]
    reg_abbrev = regulator.get("abbreviation", "REG").replace(" ", "_")[:6]
    action = extracted_enforcement.get("enforcement_action", {})
    date = action.get("notice_date", datetime.now().strftime("%Y-%m-%d"))
    return f"gap_analysis_{reg_abbrev}_{abbrev}_{date}.xlsx"
